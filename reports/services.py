# reports/services.py
"""
Service layer for the Reports app.

This module contains all business logic for generating reports
and exporting them to PDF and Excel. All functions are pure data
fetchers or exporters – they do not handle HTTP directly.
"""

import io
from datetime import datetime
from decimal import Decimal

from django.db.models import Sum, F
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Import models from your apps
from sales.models import Sale, SaleItem, Product
from expenses.models import Expense
from customers.models import Customer


# =============================================================================
# 1. REPORT DATA FUNCTIONS
# =============================================================================

def get_business_kpis(org, start_date, end_date):
    """
    Return key business KPIs for a given date range:
    - total sales revenue
    - total expenses
    - gross profit (sales - expenses)
    - cash received (all sales except credit)
    - customer credit (total credit sales)
    """
    sales = Sale.objects.filter(
        organization=org,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')

    expenses = Expense.objects.filter(
        organization=org,
        date__gte=start_date,
        date__lte=end_date
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

    cash = Sale.objects.filter(
        organization=org,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).exclude(payment_method='credit').aggregate(total=Sum('total'))['total'] or Decimal('0.00')

    credit = Sale.objects.filter(
        organization=org,
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
        payment_method='credit'
    ).aggregate(total=Sum('total'))['total'] or Decimal('0.00')

    return {
        'total_sales': sales,
        'total_expenses': expenses,
        'gross_profit': sales - expenses,
        'cash_received': cash,
        'customer_credit': credit,
    }


def get_top_products(org, start_date, end_date, limit=5):
    """
    Return the top `limit` products by revenue in the given date range.
    Each dict contains: product__name, total_qty, total_revenue.
    """
    return SaleItem.objects.filter(
        sale__organization=org,
        sale__created_at__date__gte=start_date,
        sale__created_at__date__lte=end_date
    ).values('product__name').annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum(F('quantity') * F('unit_price'))
    ).order_by('-total_revenue')[:limit]


def get_current_stock(org):
    """
    Return all products for the organization, annotated with
    `total_value` (price × stock_quantity). We use a different name
    to avoid conflict with the model's `stock_value` property.
    """
    return Product.objects.filter(organization=org).annotate(
        total_value=F('price') * F('stock_quantity')
    ).order_by('name')


def get_low_stock(org):
    """Return products whose stock is at or below the low‑stock threshold."""
    return Product.objects.filter(
        organization=org,
        stock_quantity__lte=F('low_stock_threshold')
    )


def get_customer_debt(org):
    """Return customers with outstanding balance > 0, ordered by debt."""
    return Customer.objects.filter(
        organization=org,
        outstanding_balance__gt=0
    ).order_by('-outstanding_balance')


# =============================================================================
# 2. EXPORT FUNCTIONS (PDF & Excel)
# =============================================================================

BRAND_COLOR = colors.HexColor('#4154f1')
BRAND_DARK = colors.HexColor('#2c3aad')
TEXT_MUTED = colors.HexColor('#6c757d')
ROW_ALT = colors.HexColor('#f4f6fc')
BORDER_LIGHT = colors.HexColor('#e3e6f0')
KPI_BG = colors.HexColor('#f4f6fc')

COMPANY_NAME = "SME Business OS"


def _pdf_header_footer(canvas_obj, doc, subtitle_text):
    """Draws a repeating colored header band and footer (page number, timestamp) on every page."""
    canvas_obj.saveState()
    page_width, page_height = A4

    # --- Header band ---
    band_height = 0.55 * inch
    canvas_obj.setFillColor(BRAND_COLOR)
    canvas_obj.rect(0, page_height - band_height, page_width, band_height, stroke=0, fill=1)

    canvas_obj.setFillColor(colors.whitesmoke)
    canvas_obj.setFont('Helvetica-Bold', 11)
    canvas_obj.drawString(30, page_height - band_height + 18, COMPANY_NAME)

    canvas_obj.setFont('Helvetica', 9)
    canvas_obj.drawRightString(page_width - 30, page_height - band_height + 18, subtitle_text or "")

    # --- Footer ---
    canvas_obj.setFillColor(TEXT_MUTED)
    canvas_obj.setFont('Helvetica', 8)
    canvas_obj.drawString(30, 22, f"Generated on {datetime.now().strftime('%d %b %Y, %H:%M')}")
    canvas_obj.drawRightString(page_width - 30, 22, f"Page {doc.page}")
    canvas_obj.setStrokeColor(BORDER_LIGHT)
    canvas_obj.line(30, 34, page_width - 30, 34)

    canvas_obj.restoreState()


def generate_pdf_response(filename, title, subtitle, headers, rows, summary_dict=None):
    """
    Generate a professionally styled PDF report and return an HttpResponse.
    - filename: name of the file (without extension)
    - title: report title (shown large, below the header band)
    - subtitle: date range or period
    - headers: list of column headers
    - rows: list of lists (data rows)
    - summary_dict: optional dict of KPIs to display as summary cards
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=30, leftMargin=30,
        topMargin=0.55 * inch + 24, bottomMargin=45,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='ReportTitle',
        parent=styles['Heading1'],
        alignment=TA_LEFT,
        fontSize=19,
        leading=23,
        spaceAfter=2,
        textColor=colors.HexColor('#212529'),
        fontName='Helvetica-Bold',
    )
    subtitle_style = ParagraphStyle(
        name='Subtitle',
        parent=styles['Normal'],
        alignment=TA_LEFT,
        fontSize=10.5,
        textColor=TEXT_MUTED,
        spaceAfter=18,
    )
    section_style = ParagraphStyle(
        name='Section',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#212529'),
        fontName='Helvetica-Bold',
        spaceBefore=6,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        name='Body',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_LEFT,
        textColor=TEXT_MUTED,
    )
    kpi_value_style = ParagraphStyle(
        name='KPIValue',
        parent=styles['Normal'],
        fontSize=15,
        leading=18,
        fontName='Helvetica-Bold',
        textColor=BRAND_DARK,
        alignment=TA_LEFT,
    )
    kpi_label_style = ParagraphStyle(
        name='KPILabel',
        parent=styles['Normal'],
        fontSize=8.5,
        textColor=TEXT_MUTED,
        alignment=TA_LEFT,
    )

    story = []

    # Title & Subtitle
    story.append(Paragraph(title, title_style))
    if subtitle:
        story.append(Paragraph(subtitle, subtitle_style))

    # Summary KPIs as bordered cards, laid out in a single row
    if summary_dict:
        story.append(Paragraph("SUMMARY", section_style))
        card_cells = []
        for label, value in summary_dict.items():
            cell_content = [
                Paragraph(str(value), kpi_value_style),
                Paragraph(str(label).upper(), kpi_label_style),
            ]
            card_cells.append(cell_content)

        usable_width = A4[0] - 60
        col_width = usable_width / len(card_cells)
        kpi_table = Table([card_cells], colWidths=[col_width] * len(card_cells))
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), KPI_BG),
            ('BOX', (0, 0), (-1, -1), 0.75, BORDER_LIGHT),
            ('INNERGRID', (0, 0), (-1, -1), 0.75, colors.white),
            ('LEFTPADDING', (0, 0), (-1, -1), 14),
            ('RIGHTPADDING', (0, 0), (-1, -1), 14),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.28 * inch))

    # Data Table
    if rows:
        story.append(Paragraph("DETAIL", section_style))

        header_row = [Paragraph(f"<b>{h}</b>", ParagraphStyle(
            name='TH', parent=styles['Normal'], fontSize=9.5,
            textColor=colors.whitesmoke, alignment=TA_LEFT if i == 0 else TA_CENTER
        )) for i, h in enumerate(headers)]

        cell_style_left = ParagraphStyle(name='TDLeft', parent=styles['Normal'], fontSize=9.5, textColor=colors.HexColor('#212529'))
        cell_style_right = ParagraphStyle(name='TDRight', parent=styles['Normal'], fontSize=9.5, textColor=colors.HexColor('#212529'), alignment=2)  # 2 = right

        data_rows = []
        for row in rows:
            formatted = []
            for i, val in enumerate(row):
                style = cell_style_left if i == 0 else cell_style_right
                formatted.append(Paragraph(str(val), style))
            data_rows.append(formatted)

        table_data = [header_row] + data_rows
        col_widths = [(A4[0] - 60) / len(headers)] * len(headers)
        t = Table(table_data, colWidths=col_widths, repeatRows=1)

        style = [
            ('BACKGROUND', (0, 0), (-1, 0), BRAND_COLOR),
            ('TOPPADDING', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 9),
            ('TOPPADDING', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 7),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('RIGHTPADDING', (0, 0), (-1, -1), 10),
            ('LINEBELOW', (0, 0), (-1, 0), 0.75, BRAND_DARK),
            ('LINEBELOW', (0, 1), (-1, -2), 0.5, BORDER_LIGHT),
            ('BOX', (0, 0), (-1, -1), 0.75, BORDER_LIGHT),
        ]
        # Zebra striping for readability
        for row_idx in range(1, len(table_data)):
            if row_idx % 2 == 0:
                style.append(('BACKGROUND', (0, row_idx), (-1, row_idx), ROW_ALT))
        t.setStyle(TableStyle(style))
        story.append(t)
    else:
        story.append(Paragraph("No data available for this period.", body_style))

    def _draw_page(canvas_obj, doc_obj):
        _pdf_header_footer(canvas_obj, doc_obj, subtitle)

    doc.build(story, onFirstPage=_draw_page, onLaterPages=_draw_page)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    response.write(pdf)
    return response


def generate_excel_response(filename, title, headers, rows, summary_dict=None):
    """
    Generate an Excel file (.xlsx) and return an HttpResponse.
    - filename: name of the file (without extension)
    - title: bold title at the top (also used as sheet name)
    - headers: list of column headers
    - rows: list of lists (data rows)
    - summary_dict: optional dict of KPIs to display at the top
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    # Define styles
    bold_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4154f1", end_color="4154f1", fill_type="solid")
    kpi_font = Font(bold=True, size=12, color="4154f1")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    current_row = 1

    # 1. Title and date
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    title_cell = ws.cell(row=current_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=16, color="4154f1")
    title_cell.alignment = Alignment(horizontal='center')
    current_row += 1

    subtitle = f"Generated on {datetime.now().strftime('%d %b %Y, %H:%M')}"
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    sub_cell = ws.cell(row=current_row, column=1, value=subtitle)
    sub_cell.font = Font(size=10, color="808080")
    sub_cell.alignment = Alignment(horizontal='center')
    current_row += 2

    # 2. Summary KPIs (if provided)
    if summary_dict:
        for label, value in summary_dict.items():
            ws.cell(row=current_row, column=1, value=label)
            val_cell = ws.cell(row=current_row, column=2, value=value)
            val_cell.font = kpi_font
            current_row += 1
        current_row += 1

    # 3. Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    current_row += 1

    # 4. Data rows
    for row_data in rows:
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
            cell.border = border
        current_row += 1

    # 5. Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = min(adjusted_width, 40)

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    return response